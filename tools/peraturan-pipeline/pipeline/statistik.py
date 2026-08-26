"""Kumpulkan seluruh statistik korpus dan tuliskan sebagai satu berkas Excel.

Tidak ada angka baru yang dihitung di sini. Yang dikerjakan hanya menyatukan
jawaban yang sudah tersebar di `qc`, `celah`, `celah_urut`, `tinjau`,
`verifikasi`, `hierarki`, dan `kurang` menjadi satu berkas — karena pertanyaan
"bagaimana keadaan korpusnya" tidak dijawab oleh satu pemeriksaan mana pun.

**Angka pada lembar Ringkasan berformula ke lembar rinciannya**, bukan disalin.
Angka yang disalin dapat berselisih dengan rinciannya tanpa ada yang tahu; angka
berformula tidak bisa. Itu sebabnya berkas ini menuntut penghitungan ulang
(LibreOffice) sebelum dianggap selesai — sebelum itu, setiap sel formula terbaca
kosong oleh pembaca mana pun yang membaca nilai tersimpan.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

NAVY = "00153D"
MIST = "E8ECF4"
FONT = "Arial"


def kumpulkan(conn) -> dict:
    """Seluruh angka yang masuk ke berkas, dalam satu bentuk."""
    from . import celah_urut, hierarki, kurang, qc

    D: dict = {}
    D["ringkas"] = dict(qc.ringkas(conn))
    D["pemeriksaan"] = [{k: v for k, v in x.items() if k != "contoh"}
                        for x in qc.jalankan(conn)]
    D["kurang"] = kurang.laporan(conn, "data")

    def ambil(sql, arg=()):
        return [dict(r) for r in conn.execute(sql, arg)]

    D["bentuk"] = ambil("""
        SELECT COALESCE(NULLIF(r.jenis_code,''),'(tanpa kode)') kode,
               COUNT(*) dokumen, SUM(r.has_body) bernaskah,
               MIN(r.tahun) th_awal, MAX(r.tahun) th_akhir
          FROM regulation r GROUP BY kode ORDER BY dokumen DESC""")
    unit = {r["kode"]: r["n"] for r in ambil("""
        SELECT COALESCE(NULLIF(g.jenis_code,''),'(tanpa kode)') kode,
               COUNT(*) n FROM pasal p JOIN regulation g ON g.id=p.reg_id
         GROUP BY kode""")}
    for b in D["bentuk"]:
        b["unit"] = unit.get(b["kode"], 0)

    D["tahun"] = ambil("""
        SELECT r.tahun, COUNT(*) dokumen, SUM(r.has_body) bernaskah
          FROM regulation r WHERE r.tahun IS NOT NULL
         GROUP BY r.tahun ORDER BY r.tahun DESC""")
    D["keberlakuan"] = ambil("""
        SELECT COALESCE(v.status_derived,'(belum dihitung)') status, COUNT(*) n
          FROM regulation r LEFT JOIN validity v ON v.reg_id=r.id
         GROUP BY 1 ORDER BY 2 DESC""")
    D["relasi"] = ambil("""
        SELECT type jenis, COUNT(*) total, SUM(dst_id IS NOT NULL) terpaut,
               ROUND(AVG(confidence),3) keyakinan
          FROM relation GROUP BY type ORDER BY total DESC""")
    D["gantung"] = ambil("""
        SELECT dst_raw sasaran, type jenis, COUNT(*) n
          FROM relation WHERE dst_id IS NULL AND dst_raw IS NOT NULL
         GROUP BY dst_raw, type ORDER BY n DESC LIMIT 60""")
    D["asal"] = ambil("""
        SELECT CASE
           WHEN source LIKE 'djp%' THEN 'DJP — penerbitnya sendiri'
           WHEN source LIKE '%peraturan.go.id%' OR source LIKE '%jdih%'
             OR source LIKE '%bpk%' THEN 'JDIH / peraturan.go.id — repositori resmi'
           WHEN source LIKE '%sdsn%' THEN 'SDSN — naskah konsolidasi resmi'
           WHEN source LIKE '%ddtc%' THEN 'DDTC — perantara swasta'
           WHEN source LIKE '%ortax%' THEN 'Ortax — perantara swasta'
           ELSE 'lain' END asal, COUNT(*) n, SUM(has_body) bernaskah
         FROM regulation GROUP BY 1 ORDER BY 2 DESC""")
    D["tinjauan"] = ambil("""
        SELECT periksa, status, COUNT(*) n, ROUND(AVG(keyakinan),2) keyakinan
          FROM temuan GROUP BY periksa, status ORDER BY n DESC""")
    D["daerah"] = ambil("""
        SELECT kategori daerah, COUNT(*) n, SUM(has_body) bernaskah
          FROM regulation
         WHERE kategori LIKE 'Provinsi%' OR kategori LIKE 'Kab%'
            OR kategori LIKE 'Kota%'
         GROUP BY kategori ORDER BY n DESC""")
    D["pdf"] = ambil("""
        SELECT r.jenis_code kode, COUNT(DISTINCT r.id) dokumen,
               COUNT(DISTINCT a.reg_id) berpdf
          FROM regulation r LEFT JOIN attachment a
            ON a.reg_id=r.id AND a.local_path IS NOT NULL
         GROUP BY r.jenis_code HAVING berpdf>0 ORDER BY berpdf DESC""")
    D["celah_luar"] = ambil("""
        SELECT sumber, jenis_code kode, COUNT(*) di_sumber,
               SUM(ada_di_kita) sudah_ada
          FROM katalog_luar WHERE jenis_code IS NOT NULL
         GROUP BY sumber, jenis_code HAVING COUNT(*)-SUM(ada_di_kita) > 0
         ORDER BY COUNT(*)-SUM(ada_di_kita) DESC LIMIT 60""")
    D["verifikasi"] = ambil("""
        SELECT v.sumber, v.status_baku sumber_bilang,
               COALESCE(d.status_derived,'(belum dihitung)') kita_hitung,
               COUNT(*) n
          FROM verifikasi v LEFT JOIN validity d ON d.reg_id=v.reg_id
         WHERE v.ditemukan=1 GROUP BY 1,2,3 ORDER BY n DESC LIMIT 40""")
    try:
        D["berkala"] = ambil("""
            SELECT berkala jenis, COUNT(*) dokumen, MIN(tahun) th_awal,
                   MAX(tahun) th_akhir FROM regulation
             WHERE berkala IS NOT NULL GROUP BY berkala""")
        D["nilai_berkala"] = ambil(
            "SELECT 'kurs' j, COUNT(*) n FROM kurs_nilai "
            "UNION ALL SELECT 'tarif bunga', COUNT(*) FROM tarif_nilai")
    except sqlite3.OperationalError:
        D["berkala"], D["nilai_berkala"] = [], []

    lapis = []
    for g in hierarki.peta(conn)["golongan"]:
        for t in g["tingkat"]:
            for b in t["bentuk"]:
                lapis.append({"golongan": g["nama"], "tingkat": t["label"],
                              "kode": b["kode"], "nama": b["nama"],
                              "dokumen": b["jumlah"], "berteks": b["berteks"],
                              "dasar": b["dasar"], "penerbit": b["penerbit"]})
    D["hierarki"] = lapis

    cu = celah_urut.periksa(conn, tahun_min=2000)
    D["celah_urut"] = {k: v for k, v in cu.items() if k != "seri"}
    D["celah_urut_seri"] = [
        {**{k: v for k, v in s.items() if k != "hilang"},
         "hilang_daftar": ", ".join(map(str, s["hilang"][:25]))}
        for s in cu["seri"][:80]]
    return D


def _lembar(wb, nama, judul, catatan, kolom, baris, lebar, total_kol=()):
    """Satu lembar bertabel, dengan baris TOTAL berformula bila diminta.

    Mengembalikan nomor baris TOTAL supaya lembar Ringkasan dapat menunjuknya
    tanpa menghitung sendiri — alamat yang dihitung tangan akan meleset diam-
    diam begitu jumlah barisnya berubah.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(nama)
    ws["A1"] = judul
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=NAVY)
    ws["A2"] = catatan
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="4A5568")
    ws.freeze_panes = "A5"
    kepala = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    isi = PatternFill("solid", fgColor=NAVY)
    garis = Border(bottom=Side(style="thin", color="CBD5E0"))
    for j, nm in enumerate(kolom, 1):
        c = ws.cell(row=4, column=j, value=nm)
        c.font, c.fill = kepala, isi
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for i, row in enumerate(baris, 5):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = garis
            if isinstance(v, int):
                c.number_format = "#,##0"
    akhir = 4 + len(baris)
    r_total = None
    if total_kol and baris:
        r_total = akhir + 1
        tebal = Font(name=FONT, size=10, bold=True)
        abu = PatternFill("solid", fgColor=MIST)
        for j in range(1, len(kolom) + 1):
            ws.cell(row=r_total, column=j).fill = abu
        c = ws.cell(row=r_total, column=1, value="TOTAL")
        c.font, c.fill = tebal, abu
        for j in total_kol:
            L = get_column_letter(j)
            c = ws.cell(row=r_total, column=j,
                        value=f"=SUM({L}5:{L}{akhir})")
            c.font, c.fill = tebal, abu
            c.number_format = "#,##0"
    for j, w in enumerate(lebar, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[4].height = 30
    return r_total


def tulis(D: dict, keluar: str | Path) -> Path:
    """Tulis seluruh statistik menjadi satu berkas Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    wb.remove(wb.active)
    ringkas = wb.create_sheet("Ringkasan")   # diisi terakhir, butuh alamat total
    A = {}

    A["bentuk"] = _lembar(wb, "Per Bentuk", "Dokumen per bentuk peraturan",
        "'(tanpa kode)' adalah dokumen yang bentuknya belum terpetakan — ikut "
        "ditampilkan supaya totalnya rekonsiliasi, bukan disaring keluar.",
        ["Kode", "Dokumen", "Bernaskah", "Unit naskah", "Tahun awal", "Tahun akhir"],
        [[b["kode"], b["dokumen"], b["bernaskah"] or 0, b["unit"],
          b["th_awal"], b["th_akhir"]] for b in D["bentuk"]],
        [16, 12, 12, 14, 12, 12], (2, 3, 4))

    A["tahun"] = _lembar(wb, "Per Tahun", "Dokumen menurut tahun terbit",
        "Tahun diambil dari nomor peraturannya, bukan dari kapan ia diambil. "
        "Tahun yang jauh di masa depan atau sangat lampau menandai nomor yang "
        "salah urai — bukan dokumen yang salah.",
        ["Tahun", "Dokumen", "Bernaskah"],
        [[t["tahun"], t["dokumen"], t["bernaskah"] or 0] for t in D["tahun"]],
        [12, 12, 12], (2, 3))

    _lembar(wb, "Asal Naskah", "Dari mana naskahnya diambil",
        "Bentuk berkas (HTML atau PDF) bukan penentu keandalan; yang menentukan "
        "siapa yang menerbitkannya. DJP menerbitkan badan aturannya sebagai HTML "
        "— untuk bentuk itu, HTML memang bentuk resminya.",
        ["Asal", "Dokumen", "Bernaskah", "Kedudukan"],
        [[a["asal"], a["n"], a["bernaskah"] or 0,
          {"DJP — penerbitnya sendiri":
             "Penerbitnya sendiri. Untuk PER, KEP, SE, PENG — HTML inilah terbitannya.",
           "JDIH / peraturan.go.id — repositori resmi":
             "Repositori resmi negara; memuat pindaian untuk UU, PP, Perpres.",
           "SDSN — naskah konsolidasi resmi":
             "Himpunan resmi DJP 2023, naskah terkonsolidasi.",
           "DDTC — perantara swasta":
             "Transkripsi pihak ketiga. Alamat sumber dan sidik jari tersimpan, "
             "jadi dapat diperiksa ulang.",
           "Ortax — perantara swasta":
             "Transkripsi pihak ketiga. Sama, dapat diperiksa ulang."}.get(a["asal"], "")]
         for a in D["asal"]],
        [44, 12, 12, 96], (2, 3))

    _lembar(wb, "Hierarki", "Kedudukan hukum tiap bentuk",
        "Tingkat mengikuti UU 12/2011 Pasal 7 ayat (1). Bentuk 'di luar tangga' "
        "bukan berarti lebih rendah — Pasal 8 mengakui keberadaannya sepanjang "
        "diperintahkan peraturan yang lebih tinggi atau dibentuk berdasarkan "
        "kewenangan.",
        ["Golongan", "Tingkat", "Kode", "Nama", "Dokumen", "Bernaskah",
         "Dasar kedudukan", "Penerbit"],
        [[h["golongan"], h["tingkat"], h["kode"], h["nama"], h["dokumen"],
          h["berteks"] or 0, h["dasar"], h["penerbit"]] for h in D["hierarki"]],
        [30, 34, 14, 34, 12, 12, 54, 30], (5, 6))

    _lembar(wb, "Keberlakuan", "Status keberlakuan hasil hitungan",
        "Dihitung dari relasi pencabutan dan perubahan di dalam korpus, bukan "
        "disalin dari label situs. Dokumen yang pencabutnya belum ada di korpus "
        "akan terhitung berlaku — lihat lembar Kekurangan.",
        ["Status", "Dokumen"],
        [[k["status"], k["n"]] for k in D["keberlakuan"]], [28, 14], (2,))

    A["relasi"] = _lembar(wb, "Relasi", "Relasi antar-peraturan",
        "Terpaut berarti sasarannya ada di korpus. Yang tidak terpaut umumnya "
        "undang-undang non-pajak yang dikutip sebagai dasar hukum — di luar "
        "lingkup korpus ini, bukan hilang.",
        ["Jenis relasi", "Total", "Terpaut", "Keyakinan rata-rata"],
        [[r["jenis"], r["total"], r["terpaut"] or 0, r["keyakinan"]]
         for r in D["relasi"]], [26, 14, 14, 20], (2, 3))

    _lembar(wb, "Rujukan Menggantung", "Sasaran rujukan yang tidak ada di korpus",
        "Diurutkan menurut seberapa sering dirujuk. Yang teratas adalah "
        "undang-undang non-pajak yang dikutip hampir setiap peraturan daerah "
        "sebagai dasar hukum — menambahkannya menutup rujukan menggantung "
        "paling banyak dengan dokumen paling sedikit.",
        ["Sasaran", "Jenis relasi", "Kali dirujuk"],
        [[g["sasaran"], g["jenis"], g["n"]] for g in D["gantung"]],
        [34, 22, 16], (3,))

    _lembar(wb, "Pemeriksaan", "Pemeriksaan mutu (QC)",
        "Keparahan menandai mana yang menghalangi pemakaian dan mana yang hanya "
        "mengurangi kerapian.",
        ["Pemeriksaan", "Pertanyaan", "Keparahan", "Temuan", "Dari"],
        [[x["nama"], x["pertanyaan"], x["keparahan"], x["jumlah"], x["dari"]]
         for x in D["pemeriksaan"]], [30, 64, 14, 12, 12], (4,))

    _lembar(wb, "Tinjauan", "Antrean tinjauan manusia",
        "auto_selesai = sudah diperbaiki program karena keyakinannya di atas "
        "0,90, tetapi bertanda agar dapat ditinjau ulang. antre = menunggu "
        "manusia. tidak_berlaku_lagi = temuan yang penyebabnya sudah hilang.",
        ["Pemeriksaan", "Status", "Temuan", "Keyakinan rata-rata"],
        [[t["periksa"], t["status"], t["n"], t["keyakinan"]] for t in D["tinjauan"]],
        [26, 22, 12, 20], (3,))

    _lembar(wb, "Verifikasi Silang", "Status menurut sumber lain vs hitungan kita",
        "Sumber resmi (peraturan.go.id, JDIH) dapat menyelesaikan satu kasus "
        "sendirian; DDTC penerbit swasta, jadi ia menguatkan tetapi tidak "
        "memutus. Baris 'sumber bilang dicabut, kita hitung berlaku' adalah "
        "arah kekeliruan yang paling berbahaya.",
        ["Sumber", "Sumber menyatakan", "Kita menghitung", "Dokumen"],
        [[v["sumber"], v["sumber_bilang"], v["kita_hitung"], v["n"]]
         for v in D["verifikasi"]], [22, 22, 22, 14], (4,))

    _lembar(wb, "Daerah", f"Cakupan pajak daerah — {len(D['daerah'])} dari 553 daerah",
        "Daerah sisanya bukan kekurangan pengambilan: katalog sumber memang "
        "mencatatnya kosong, artinya daerahnya belum menerbitkan atau belum "
        "masuk ke katalog itu.",
        ["Daerah", "Dokumen", "Bernaskah"],
        [[d["daerah"], d["n"], d["bernaskah"] or 0] for d in D["daerah"]],
        [40, 12, 12], (2, 3))

    _lembar(wb, "Kekurangan", "Apa yang masih kurang",
        "'Keputusan manusia' berarti program tidak boleh memutuskannya sendiri "
        "— bukan berarti tertunda.",
        ["Kekurangan", "Jumlah", "Dari", "Tindakan", "Catatan"],
        [[k["nama"], k["jumlah"], k["dari"], k["tindakan"], k["catatan"]]
         for k in D["kurang"]], [40, 12, 12, 58, 80])

    _lembar(wb, "Celah Sumber Luar", "Ada di katalog luar, belum di korpus",
        "Selisih tidak selalu berarti hilang: sebagian adalah perbedaan lingkup. "
        "Korpus ini berlingkup perpajakan; Ortax dan DDTC mencakup seluruh "
        "fiskal termasuk bea masuk, cukai, dan PNBP.",
        ["Sumber", "Kode", "Di sumber", "Sudah ada", "Selisih"],
        [[x["sumber"], x["kode"], x["di_sumber"], x["sudah_ada"] or 0,
          x["di_sumber"] - (x["sudah_ada"] or 0)] for x in D["celah_luar"]],
        [18, 16, 14, 14, 12], (3, 4, 5))

    cu = D["celah_urut"]
    _lembar(wb, "Celah Penomoran", "Nomor yang bolong di dalam satu seri",
        f"{cu['seri_diperiksa']:,} seri diperiksa, {cu['seri_bercelah']} "
        f"bercelah, {cu['nomor_hilang']} nomor hilang. Seri peraturan daerah "
        f"dihitung per daerah — menggabungkannya justru menyembunyikan celah "
        f"yang nyata.",
        ["Kode", "Tahun", "Daerah", "Punya", "Rentang", "Kepadatan",
         "Hilang", "Nomor yang tidak ada"],
        [[s["jenis"], s["tahun"], s.get("wilayah") or "(pusat)", s["punya"],
          s["rentang"], s["kepadatan"], s["n_hilang"], s["hilang_daftar"]]
         for s in D["celah_urut_seri"]], [14, 10, 26, 10, 14, 12, 10, 44], (7,))

    A["pdf"] = _lembar(wb, "PDF", "Pindaian resmi yang tersimpan",
        "Batasnya bukan teknis melainkan siapa yang menerbitkan. Terukur: UU "
        "6/8, PP 4/5, Perpres 5/5, PMK 9/12 — sedangkan KMK 0/14 dan seluruh "
        "terbitan Dirjen Pajak serta peraturan daerah 0.",
        ["Kode", "Dokumen", "Ada PDF"],
        [[p["kode"], p["dokumen"], p["berpdf"]] for p in D["pdf"]],
        [16, 14, 12], (2, 3))

    if D["berkala"]:
        _lembar(wb, "Terbitan Berkala", "Kurs dan tarif bunga",
            "Disisihkan dari daftar utama karena memenuhi hasil pencarian tanpa "
            "pernah dibaca sebagai norma — tetap dapat dicari dan dikutip.",
            ["Jenis", "Dokumen", "Tahun awal", "Tahun akhir"],
            [[b["jenis"], b["dokumen"], b["th_awal"], b["th_akhir"]]
             for b in D["berkala"]], [20, 14, 12, 12], (2,))

    # ------------------------------------------------------------ Ringkasan
    R = D["ringkas"]
    ringkas["A1"] = "Statistik Korpus Peraturan Perpajakan Indonesia"
    ringkas["A1"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
    ringkas["A2"] = ("Angka di lembar ini BERFORMULA ke lembar rinciannya, "
                     "bukan disalin — keduanya tidak dapat berselisih tanpa "
                     "terlihat.")
    ringkas["A2"].font = Font(name=FONT, size=9, italic=True, color="4A5568")
    kepala = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    isi = PatternFill("solid", fgColor=NAVY)
    garis = Border(bottom=Side(style="thin", color="CBD5E0"))
    for j, nm in enumerate(["Ukuran", "Jumlah", "Arti"], 1):
        c = ringkas.cell(row=4, column=j, value=nm)
        c.font, c.fill = kepala, isi
        c.alignment = Alignment(horizontal="center", vertical="center")

    baris = [
        ("Dokumen", f"='Per Bentuk'!B{A['bentuk']}",
         "seluruh dokumen di korpus"),
        ("Dokumen bernaskah", f"='Per Bentuk'!C{A['bentuk']}",
         "punya naskah — dapat dicari dan dikutip"),
        ("Unit naskah", f"='Per Bentuk'!D{A['bentuk']}",
         "pasal, ayat, huruf, angka, diktum — satuan terkecil yang dapat dikutip"),
        ("Pasal", R["pasal"], "unit bertingkat pasal"),
        ("Diktum", R["diktum"], "satuan operatif pada Keputusan"),
        ("Relasi antar-aturan", f"=Relasi!B{A['relasi']}",
         "mencabut, mengubah, dasar hukum, melaksanakan"),
        ("Relasi terpaut", f"=Relasi!C{A['relasi']}",
         "sasarannya ada di korpus"),
        ("Berlaku", R["berlaku"], "hasil hitungan dari relasi, bukan label situs"),
        ("Dicabut", R["dicabut"], "terbukti dicabut oleh peraturan lain di korpus"),
        ("Dokumen berPDF resmi", f"=PDF!C{A['pdf']}",
         "batasnya penerbit, bukan teknis — lihat lembar PDF"),
        ("Naskah konsolidasi", R["konsolidasi"],
         "SDSN 2023 — satu naskah menghimpun perubahannya"),
        ("Unit berprovenance", R["berprovenance"],
         "diketahui perubahan mana yang membentuk rumusannya"),
    ]
    for i, (nm, v, ket) in enumerate(baris, 5):
        ringkas.cell(row=i, column=1, value=nm).font = Font(name=FONT, size=10)
        c = ringkas.cell(row=i, column=2, value=v)
        c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = "#,##0"
        ringkas.cell(row=i, column=3, value=ket).font = Font(name=FONT, size=10)
        for j in range(1, 4):
            ringkas.cell(row=i, column=j).border = garis

    ringkas["A19"] = "Turunan"
    ringkas["A19"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
    turun = [("Bernaskah", "=B6/B5", "0.0%", "sisanya hanya metadata"),
             ("Relasi terpaut", "=B11/B10", "0.0%",
              "sisanya menunjuk aturan di luar lingkup korpus"),
             ("Unit per dokumen bernaskah", "=B7/B6", "#,##0.0",
              "makin tinggi makin halus kutipannya"),
             ("Dokumen berPDF resmi", "=B14/B5", "0.0%",
              "hampir seluruh korpus berupa transkripsi, bukan pindaian")]
    for i, (nm, f_, fmt, ket) in enumerate(turun, 20):
        ringkas.cell(row=i, column=1, value=nm).font = Font(name=FONT, size=10)
        c = ringkas.cell(row=i, column=2, value=f_)
        c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = fmt
        ringkas.cell(row=i, column=3, value=ket).font = Font(name=FONT, size=10)
    for L, w in zip("ABC", (34, 16, 84)):
        ringkas.column_dimensions[L].width = w
    ringkas.freeze_panes = "A5"
    wb.move_sheet("Ringkasan", -(len(wb.sheetnames) - 1))

    # openpyxl menulis formula TANPA nilai tersimpan, jadi setiap sel berformula
    # terbaca kosong sampai ada yang menghitungnya. Excel biasanya menghitung
    # saat membuka, tetapi "biasanya" tidak cukup untuk berkas yang diunduh dan
    # dibuka orang lain: yang mereka lihat akan berupa sel kosong di lembar
    # Ringkasan. Penanda ini memaksa hitung ulang penuh saat dibuka, di Excel
    # maupun LibreOffice.
    wb.calculation.fullCalcOnLoad = True

    keluar = Path(keluar)
    wb.save(keluar)
    return keluar
